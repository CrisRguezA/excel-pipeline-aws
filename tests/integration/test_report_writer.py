import pytest
import pandas as pd
from datetime import datetime
import openpyxl

from excel_pipeline.stage_3_formatting.report_writer import write_report


_REQUIRED_COLS = [
    "id_venta", "cliente", "fecha_venta", "producto", "tipo_madera",
    "certificacion", "cantidad_m3", "precio_m3", "importe", "estado",
    "comercial", "pais",
]

_EXPECTED_INFO_LABELS = [
    "Proyecto",
    "Archivo fuente",
    "Fecha generación",
    "Total filas",
    "Ventas cerradas",
    "Ventas certificadas",
    "Columnas",
    "Pipeline",
]


@pytest.fixture(scope="module")
def sample_df():
    return pd.DataFrame({
        "id_venta":      ["0001", "0002", "0003"],
        "cliente":       ["ACME", "Beta Corp", "Gamma"],
        "fecha_venta":   [datetime(2026, 1, 5), datetime(2026, 1, 6), datetime(2026, 1, 7)],
        "producto":      ["Tablero", "Viga", "Panel"],
        "tipo_madera":   ["Pino", "Roble", "Eucalipto"],
        "certificacion": ["FSC", "PEFC", "Sin certificación"],
        "cantidad_m3":   [10.5, 5.0, 3.5],
        "precio_m3":     [150.0, 200.0, 120.0],
        "importe":       [1575.0, 1000.0, 420.0],
        "estado":        ["Cerrado", "Pendiente", "Cancelado"],
        "comercial":     ["Ana", "Luis", "María"],
        "pais":          ["España", "Francia", "Portugal"],
    })


@pytest.fixture(scope="module")
def report_path(sample_df, tmp_path_factory):
    out = tmp_path_factory.mktemp("stage3")
    return write_report(sample_df, out)


class TestWriteReport:
    def test_file_created(self, report_path):
        assert report_path.exists()

    def test_sheet_names(self, report_path):
        wb = openpyxl.load_workbook(report_path)
        assert wb.sheetnames == ["Weekly_Report", "Report_Info"]

    def test_info_labels(self, report_path):
        wb = openpyxl.load_workbook(report_path)
        ws = wb["Report_Info"]
        labels = [ws.cell(row=i, column=1).value for i in range(1, 9)]
        assert labels == _EXPECTED_INFO_LABELS

    def test_info_counts(self, report_path):
        wb = openpyxl.load_workbook(report_path)
        ws = wb["Report_Info"]
        # sample_df: 1 "Cerrado" row, 2 certified rows (FSC + PEFC)
        assert ws.cell(row=5, column=2).value == 1   # Ventas cerradas
        assert ws.cell(row=6, column=2).value == 2   # Ventas certificadas

    def test_totals_row_values(self, report_path):
        wb = openpyxl.load_workbook(report_path)
        ws = wb["Weekly_Report"]
        # title=row1, header=row2, data=rows3-5, totals=row6
        totals_row   = 6
        importe_col  = _REQUIRED_COLS.index("importe") + 1
        cantidad_col = _REQUIRED_COLS.index("cantidad_m3") + 1
        assert ws.cell(row=totals_row, column=importe_col).value  == pytest.approx(2995.0)
        assert ws.cell(row=totals_row, column=cantidad_col).value == pytest.approx(19.0)

    def test_totals_row_others_blank(self, report_path):
        wb = openpyxl.load_workbook(report_path)
        ws = wb["Weekly_Report"]
        totals_row = 6
        for col_name in _REQUIRED_COLS:
            if col_name in ("importe", "cantidad_m3"):
                continue
            col_idx = _REQUIRED_COLS.index(col_name) + 1
            assert ws.cell(row=totals_row, column=col_idx).value is None, \
                f"Expected blank in column '{col_name}' at totals row"

    def test_id_venta_preserves_leading_zeros(self, report_path):
        wb = openpyxl.load_workbook(report_path)
        ws = wb["Weekly_Report"]
        # first data row is row 3, id_venta is column 1
        assert ws.cell(row=3, column=1).value == "0001"
